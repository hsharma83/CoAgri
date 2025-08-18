# farm/views.py
from datetime import timedelta, datetime
from decimal import Decimal

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.views.generic import ListView, DetailView, CreateView, FormView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum, F, ExpressionWrapper, DecimalField
from django.http import JsonResponse, HttpResponse
from .models import FarmPlan, Expense, DailyActivity, Plot, FarmPlanStep, Farmland, CashIn, Sale, InventoryMovement, Stock, Item, Crop, CropProduce
from .forms import FarmPlanForm, FarmPlanStepFormSet, SingleFarmPlanStepForm, DailyActivityFormSet, ExpenseForm, \
    CashInForm, SaleForm, InventoryMovementForm, ItemForm, CropForm
import openpyxl
from openpyxl.styles import Font, PatternFill


# A mixin for role-based queryset filtering
class RoleBasedQuerysetMixin:
    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()

        if user.is_superuser or (
                hasattr(user, 'staff') and user.staff.role.name in ['ADMIN', 'OPERATION_DIRECTOR', 'AUDITOR']):
            return queryset

        if hasattr(user, 'staff'):
            staff_profile = user.staff
            if staff_profile.role.name == 'ZONE_MANAGER':
                return queryset.filter(plot__parent_farm__parent_zone=staff_profile.assigned_zone)
            if staff_profile.role.name == 'FARM_MANAGER':
                return queryset.filter(plot__parent_farm=staff_profile.assigned_farm)
            if staff_profile.role.name == 'SUPERVISOR':
                return queryset.filter(plot=staff_profile.assigned_plot)

        return queryset.none()  # Return empty if no role matches


class FarmPlanListView(LoginRequiredMixin, RoleBasedQuerysetMixin, ListView):
    model = FarmPlan
    template_name = 'farm/farm_plan_list.html'
    context_object_name = 'plans'


class FarmlandListView(LoginRequiredMixin, ListView):
    model = Farmland
    template_name = 'farm/farmland_list.html'
    context_object_name = 'farmlands'
    
    def get_queryset(self):
        user = self.request.user
        if user.is_superuser:
            return Farmland.objects.all()
        
        if hasattr(user, 'staff'):
            staff_profile = user.staff
            if staff_profile.role.name in ['ADMIN', 'OPERATION_DIRECTOR', 'AUDITOR']:
                return Farmland.objects.all()
            elif staff_profile.role.name == 'ZONE_MANAGER':
                return Farmland.objects.filter(parent_zone=staff_profile.assigned_zone)
            elif staff_profile.role.name == 'FARM_MANAGER':
                return Farmland.objects.filter(pk=staff_profile.assigned_farm.pk)
        
        return Farmland.objects.none()


class FarmlandDetailView(LoginRequiredMixin, DetailView):
    model = Farmland
    template_name = 'farm/farmland_detail.html'
    context_object_name = 'farmland'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        farmland = self.get_object()
        plots = farmland.plots.all()
        
        # Add today's activity and sowing date for each plot
        today = timezone.now().date()
        for plot in plots:
            # Get today's planned activity from active farm plan steps
            today_activity = None
            active_plan = plot.farm_plans.filter(is_active=True).first()
            if active_plan:
                days_since_start = (today - active_plan.date_of_start).days
                step = active_plan.steps.filter(
                    start_day_offset__lte=days_since_start,
                    start_day_offset__gte=days_since_start - 7  # Within last 7 days
                ).first()
                if step:
                    today_activity = step.activity.name
            plot.today_activity = today_activity
            
            # Get sowing date from daily activities
            sowing_activity = plot.daily_activities.filter(
                activity__name__icontains='sowing'
            ).order_by('date').first()
            plot.sowing_date = sowing_activity.date if sowing_activity else None
        
        context['plots'] = plots
        
        total_plot_area = sum(plot.size for plot in plots)
        context['total_plot_area'] = total_plot_area
        context['available_space'] = farmland.size - total_plot_area
        
        return context





class FarmPlanCreateView(LoginRequiredMixin, CreateView):
    model = FarmPlan
    form_class = FarmPlanForm
    template_name = 'farm/farm_plan_form.html'
    success_url = reverse_lazy('farm:plan_list')
    
    def dispatch(self, request, *args, **kwargs):
        # Restrict access for farm officers
        if hasattr(request.user, 'staff') and request.user.staff.role.name == 'FARM_MANAGER':
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied("Farm officers cannot create farm plans.")
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        """Pass the user to the form."""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['steps'] = FarmPlanStepFormSet(self.request.POST)
        else:
            data['steps'] = FarmPlanStepFormSet()
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        steps = context['steps']
        if steps.is_valid():
            self.object = form.save()
            steps.instance = self.object
            steps.save()
            return super().form_valid(form)
        else:
            return self.form_invalid(form)

class FarmPlanDetailView(LoginRequiredMixin, RoleBasedQuerysetMixin, DetailView):
    model = FarmPlan
    template_name = 'farm/farm_plan_detail.html'
    context_object_name = 'plan'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        plan = self.get_object()
        today = timezone.now().date()

        # Add the form for the modal to the context
        # If the 'form' kwarg is passed (from a failed POST), use it. Otherwise, create a new one.
        steps_with_status = []
        plan_steps = plan.steps.all().order_by('start_day_offset')

        for step in plan_steps:
            # Calculate total spent using the new cost structure
            total_spent_on_step = 0
            for entry in step.daily_entries.all():
                total_spent_on_step += entry.total_cost

            cost_variance = step.get_total_budgeted_cost() - total_spent_on_step

            cost_progress = 0
            if step.get_total_budgeted_cost() > 0:
                cost_progress = (total_spent_on_step / step.get_total_budgeted_cost()) * 100

            # ... (status logic remains the same) ...
            status = "Upcoming"
            status_color = "info"
            start_date = step.get_start_date()
            end_date = step.get_end_date()
            if today >= start_date:
                status = "Delayed Start" if not step.daily_entries.exists() else "In Progress"
                status_color = "warning" if status == "Delayed Start" else "success"
            if today > end_date and status != "Completed":
                status = "Overdue"
                status_color = "danger"

            steps_with_status.append({
                'step': step,
                'actual_cost': total_spent_on_step,  # Use the calculated total
                'cost_variance': cost_variance,
                'cost_progress': min(cost_progress, 100),
                'status': status,
                'status_color': status_color,
            })

        context['steps_with_status'] = steps_with_status

        total_budget = sum(s['step'].get_total_budgeted_cost() for s in steps_with_status)
        total_actual_cost = sum(s['actual_cost'] for s in steps_with_status)
        context['total_budget'] = total_budget
        context['total_actual_cost'] = total_actual_cost

        if total_budget > 0:
            context['budget_progress'] = min((total_actual_cost / total_budget) * 100, 100)
        else:
            context['budget_progress'] = 0
            
        # Add form for modal
        if 'add_step_form' not in kwargs:
            context['add_step_form'] = SingleFarmPlanStepForm()

        return context

    def post(self, request, *args, **kwargs):
        """Handle POST requests for adding a new FarmPlanStep."""
        # Restrict plan step creation for farm officers
        if hasattr(request.user, 'staff') and request.user.staff.role.name == 'FARM_MANAGER':
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied("Farm officers cannot create plan steps.")
            
        self.object = self.get_object()
        form = SingleFarmPlanStepForm(request.POST)

        if form.is_valid():
            new_step = form.save(commit=False)
            new_step.parent_plan = self.object
            new_step.save()
            return redirect(reverse('farm:plan_detail', kwargs={'pk': self.object.pk}))
        else:
            return self.render_to_response(self.get_context_data(add_step_form=form))


class DailyActivityCreateView(LoginRequiredMixin, CreateView):
    template_name = 'farm/activity_form.html'
    model = DailyActivity
    success_url = reverse_lazy('farm:plan_list')

    def get(self, request, *args, **kwargs):
        formset = DailyActivityFormSet(queryset=DailyActivity.objects.none())
        # Filter plots for each form based on user permissions
        for form in formset:
            self._filter_plots_for_user(form, request.user)
        return self.render_to_response({'formset': formset})
    
    def _filter_plots_for_user(self, form, user):
        """Filter plot choices based on user role and permissions"""
        if user.is_superuser:
            return  # Superuser sees all plots
        
        if hasattr(user, 'staff'):
            staff_profile = user.staff
            if staff_profile.role.name in ['ADMIN', 'OPERATION_DIRECTOR', 'AUDITOR']:
                return  # These roles see all plots
            elif staff_profile.role.name == 'ZONE_MANAGER':
                form.fields['plot'].queryset = Plot.objects.filter(
                    parent_farm__parent_zone=staff_profile.assigned_zone
                )
            elif staff_profile.role.name == 'FARM_MANAGER':
                form.fields['plot'].queryset = Plot.objects.filter(
                    parent_farm=staff_profile.assigned_farm
                )
        else:
            form.fields['plot'].queryset = Plot.objects.none()

    def post(self, request, *args, **kwargs):
        formset = DailyActivityFormSet(request.POST)
        
        # Filter plots and update queryset for parent_plan_step fields based on POST data
        for i, form in enumerate(formset):
            # Filter plots based on user permissions
            self._filter_plots_for_user(form, request.user)
            
            plot_field_name = f'form-{i}-plot'
            if plot_field_name in request.POST:
                plot_id = request.POST.get(plot_field_name)
                if plot_id:
                    try:
                        plot = Plot.objects.get(id=plot_id)
                        form.fields['parent_plan_step'].queryset = FarmPlanStep.objects.filter(
                            parent_plan__plot=plot,
                            parent_plan__is_active=True
                        )
                    except Plot.DoesNotExist:
                        pass
        
        if formset.is_valid():
            date_for_all = request.POST.get('activity_date')
            if not date_for_all:
                formset.non_form_errors().append("A date for the activities must be provided.")
                return self.render_to_response({'formset': formset})

            instances = formset.save(commit=False)
            for i, instance in enumerate(instances):
                instance.date = date_for_all
                instance.unit_of_measure = 'acres'
                instance.save()
                
                # Handle input usage
                form = formset.forms[i]
                if form.cleaned_data.get('input_used') and form.cleaned_data.get('input_item') and form.cleaned_data.get('input_quantity'):
                    from .models import ActivityItemUsage, InventoryMovement, Stock
                    
                    # Get latest stock price for the item
                    latest_stock_in = InventoryMovement.objects.filter(
                        farm=instance.plot.parent_farm,
                        item=form.cleaned_data['input_item'],
                        transaction_type='IN'
                    ).order_by('-date').first()
                    
                    cost_per_unit = latest_stock_in.price_per_unit if latest_stock_in else 0
                    
                    # Create ActivityItemUsage record
                    ActivityItemUsage.objects.create(
                        daily_activity=instance,
                        item=form.cleaned_data['input_item'],
                        quantity_used=form.cleaned_data['input_quantity'],
                        cost_per_unit=cost_per_unit
                    )
                    
                    # Create inventory movement
                    InventoryMovement.objects.create(
                        farm=instance.plot.parent_farm,
                        item=form.cleaned_data['input_item'],
                        transaction_type='OUT',
                        quantity=form.cleaned_data['input_quantity'],
                        price_per_unit=cost_per_unit,
                        date=instance.date,
                        related_activity=instance
                    )
                    
                    # Update stock levels
                    stock, created = Stock.objects.get_or_create(
                        farm=instance.plot.parent_farm,
                        item=form.cleaned_data['input_item'],
                        defaults={'quantity': 0}
                    )
                    stock.quantity -= form.cleaned_data['input_quantity']
                    stock.save()
                
                # Update the activity to trigger expense creation with item costs
                instance.save()
                    
            return redirect(self.success_url)
        
        return self.render_to_response({'formset': formset})


def get_plan_steps(request, plot_id):
    """AJAX endpoint to get plan steps for a specific plot"""
    try:
        plot = Plot.objects.get(id=plot_id)
        
        steps = FarmPlanStep.objects.filter(
            parent_plan__plot=plot,
            parent_plan__is_active=True
        ).select_related('activity', 'parent_plan')
        
        data = [{
            'id': step.id,
            'name': f"{step.activity.name} - {step.parent_plan.crop.name} (Day {step.start_day_offset})"
        } for step in steps]
        
        return JsonResponse(data, safe=False)
    except Plot.DoesNotExist:
        return JsonResponse({'error': f'Plot with id {plot_id} not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)


# A mixin for views that operate on a specific farm
class FarmContextMixin:
    def dispatch(self, request, *args, **kwargs):
        self.farm = get_object_or_404(Farmland, pk=kwargs['farm_pk'])
        return super().dispatch(request, *args, **kwargs)
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['farm'] = self.farm
        return context


# Generic Create View for our finance forms
class BaseFinanceCreateView(LoginRequiredMixin, FarmContextMixin, CreateView):
    template_name = 'farm/generic_form.html'
    form_type = 'Entry'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_type'] = self.form_type
        return context

    def form_valid(self, form):
        # Associate the object with the farm from the URL
        form.instance.farm = self.farm
        return super().form_valid(form)

    def get_success_url(self):
        # Redirect back to the finance page for the farm
        return reverse('farm:farm_finances', kwargs={'farm_pk': self.kwargs['farm_pk']})

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # For the SaleForm, pass the farm object to filter crop produce
        if self.form_class == SaleForm:
            kwargs['farm'] = get_object_or_404(Farmland, pk=self.kwargs['farm_pk'])
        return kwargs
    
    def dispatch(self, request, *args, **kwargs):
        # Check role-based permissions for SaleForm
        if self.form_class == SaleForm:
            if not request.user.is_superuser:
                if not (hasattr(request.user, 'staff') and 
                       request.user.staff.role.name in ['FARM_MANAGER', 'OPERATION_DIRECTOR', 'ADMIN']):
                    from django.core.exceptions import PermissionDenied
                    raise PermissionDenied("Only Farm Officers can add sales.")
        return super().dispatch(request, *args, **kwargs)


class ExpenseCreateView(BaseFinanceCreateView):
    form_class = ExpenseForm
    form_type = 'Expense'


class CashInCreateView(BaseFinanceCreateView):
    form_class = CashInForm
    form_type = 'Cash In'


class SaleCreateView(BaseFinanceCreateView):
    form_class = SaleForm
    form_type = 'Sale'

    def form_valid(self, form):
        return super(CreateView, self).form_valid(form)


class InventoryMovementCreateView(BaseFinanceCreateView):
    form_class = InventoryMovementForm
    form_type = 'Stock Movement'
    template_name = 'farm/generic_form.html'
    
    def form_valid(self, form):
        form.instance.farm = self.farm
        movement = form.save()
        
        # Update stock levels
        stock, created = Stock.objects.get_or_create(
            farm=self.farm,
            item=movement.item,
            defaults={'quantity': 0}
        )
        
        if movement.transaction_type == 'IN':
            stock.quantity += movement.quantity
        elif movement.transaction_type == 'OUT':
            stock.quantity -= movement.quantity
        elif movement.transaction_type == 'ADJUST':
            stock.quantity = movement.quantity
        
        stock.save()
        return super(CreateView, self).form_valid(form)


class StockRegisterView(LoginRequiredMixin, FarmContextMixin, TemplateView):
    template_name = 'farm/stock_register.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        farm = context['farm']
        
        # Get current stock levels
        stocks = Stock.objects.filter(farm=farm).select_related('item')
        
        # Get recent movements
        movements = InventoryMovement.objects.filter(
            farm=farm
        ).select_related('item').order_by('-date')[:50]
        
        context['stocks'] = stocks
        context['movements'] = movements
        
        return context


class FarmFinanceView(LoginRequiredMixin, FarmContextMixin, TemplateView):
    template_name = 'farm/farm_finance_ledger.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        farm = context['farm']

        # Handle date filtering
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')

        # Default to last 30 days if no dates are provided
        if not start_date_str:
            start_date = (datetime.now() - timedelta(days=30)).date()
        else:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()

        if not end_date_str:
            end_date = datetime.now().date()
        else:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        # Get expenses, cash ins, and sales
        expenses = Expense.objects.filter(farm=farm, date__range=[start_date, end_date])
        cash_ins = CashIn.objects.filter(farm=farm, date__range=[start_date, end_date])
        sales = Sale.objects.filter(crop_produce__farm=farm, sale_date__range=[start_date, end_date])

        # Group transactions by category
        categories = {}
        
        # Expenses
        for exp in expenses:
            category_name = exp.category.name
            if category_name not in categories:
                categories[category_name] = {
                    'type': 'Expense',
                    'transactions': [],
                    'total_amount': 0,
                    'total_units': 0
                }
            
            categories[category_name]['transactions'].append({
                'date': exp.date,
                'description': exp.description,
                'units': None,
                'unit_measure': None,
                'cost_per_unit': None,
                'amount': exp.amount
            })
            categories[category_name]['total_amount'] += exp.amount
        
        # Cash inflows
        for ci in cash_ins:
            category_name = f"Cash In - {ci.head.name}"
            if category_name not in categories:
                categories[category_name] = {
                    'type': 'Cash In',
                    'transactions': [],
                    'total_amount': 0,
                    'total_units': 0
                }
            
            categories[category_name]['transactions'].append({
                'date': ci.date,
                'description': f"From {ci.given_by}",
                'units': None,
                'unit_measure': None,
                'cost_per_unit': None,
                'amount': ci.amount
            })
            categories[category_name]['total_amount'] += ci.amount
        
        # Sales
        for sale in sales:
            category_name = f"Sales - {sale.crop_produce.crop.name}"
            if category_name not in categories:
                categories[category_name] = {
                    'type': 'Sale',
                    'transactions': [],
                    'total_amount': 0,
                    'total_units': 0
                }
            
            categories[category_name]['transactions'].append({
                'date': sale.sale_date,
                'description': f"To {sale.buyer}",
                'units': sale.quantity,
                'unit_measure': 'tons',
                'cost_per_unit': sale.price_per_ton,
                'amount': sale.total_amount
            })
            categories[category_name]['total_amount'] += sale.total_amount
            categories[category_name]['total_units'] += sale.quantity
        
        # Sort transactions within each category by date
        for category in categories.values():
            category['transactions'].sort(key=lambda x: x['date'], reverse=True)
        
        context['categories'] = categories

        # Calculate totals for the entire history of the farm for the balance
        total_cash_in = CashIn.objects.filter(farm=farm).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        total_sales = Sale.objects.filter(crop_produce__farm=farm).aggregate(
            total=Sum(F('quantity') * F('price_per_ton'))
        )['total'] or Decimal('0.00')
        total_expenses = Expense.objects.filter(farm=farm).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        context['total_inflow'] = total_cash_in + total_sales
        context['total_outflow'] = total_expenses
        context['net_balance'] = context['total_inflow'] - context['total_outflow']

        context['start_date'] = start_date.strftime('%Y-%m-%d')
        context['end_date'] = end_date.strftime('%Y-%m-%d')

        return context


class FarmFinanceExportView(LoginRequiredMixin, FarmContextMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        farm = get_object_or_404(Farmland, pk=kwargs['farm_pk'])
        
        # Get date range
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        if not start_date_str:
            start_date = (datetime.now() - timedelta(days=30)).date()
        else:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            
        if not end_date_str:
            end_date = datetime.now().date()
        else:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Get data including activity expenses
        from .models import DailyActivity
        activity_expenses = DailyActivity.objects.filter(
            plot__parent_farm=farm,
            date__range=[start_date, end_date]
        ).select_related('activity', 'plot')
        expenses = Expense.objects.filter(farm=farm, date__range=[start_date, end_date])
        cash_ins = CashIn.objects.filter(farm=farm, date__range=[start_date, end_date])
        sales = Sale.objects.filter(crop_produce__farm=farm, sale_date__range=[start_date, end_date])
        
        # Create workbook with multiple sheets
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        category_font = Font(bold=True, color="000000")
        category_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Create Activity Expenses sheet
        if activity_expenses.exists():
            ws_activities = wb.create_sheet("Activity Expenses")
            headers = ['Date', 'Activity', 'Plot', 'Machine Cost', 'Labor Cost', 'Item Cost', 'Total Cost', 'Vendor']
            for col, header in enumerate(headers, 1):
                cell = ws_activities.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            row = 2
            for activity in activity_expenses:
                if activity.total_cost > 0:
                    ws_activities.cell(row=row, column=1, value=activity.date.strftime('%Y-%m-%d'))
                    ws_activities.cell(row=row, column=2, value=activity.activity.name)
                    ws_activities.cell(row=row, column=3, value=f"Plot {activity.plot.plot_id}")
                    ws_activities.cell(row=row, column=4, value=float(activity.machine_cost))
                    ws_activities.cell(row=row, column=5, value=float(activity.labor_cost))
                    ws_activities.cell(row=row, column=6, value=float(activity.item_cost))
                    ws_activities.cell(row=row, column=7, value=float(activity.total_cost))
                    ws_activities.cell(row=row, column=8, value=activity.vendor_name or '')
                    row += 1
        
        # Create Regular Expenses sheet
        if expenses.exists():
            ws_expenses = wb.create_sheet("Regular Expenses")
            headers = ['Date', 'Category', 'Description', 'Amount (₹)']
            for col, header in enumerate(headers, 1):
                cell = ws_expenses.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            row = 2
            for exp in expenses:
                ws_expenses.cell(row=row, column=1, value=exp.date.strftime('%Y-%m-%d'))
                ws_expenses.cell(row=row, column=2, value=exp.category.name)
                ws_expenses.cell(row=row, column=3, value=exp.description)
                ws_expenses.cell(row=row, column=4, value=float(exp.amount))
                row += 1
        
        # Create Cash Inflows sheet
        if cash_ins.exists():
            ws_cashin = wb.create_sheet("Cash Inflows")
            headers = ['Date', 'Head', 'Given By', 'Amount (₹)']
            for col, header in enumerate(headers, 1):
                cell = ws_cashin.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            row = 2
            for ci in cash_ins:
                ws_cashin.cell(row=row, column=1, value=ci.date.strftime('%Y-%m-%d'))
                ws_cashin.cell(row=row, column=2, value=ci.head.name)
                ws_cashin.cell(row=row, column=3, value=ci.given_by)
                ws_cashin.cell(row=row, column=4, value=float(ci.amount))
                row += 1
        
        # Create Sales sheet
        if sales.exists():
            ws_sales = wb.create_sheet("Sales")
            headers = ['Date', 'Crop', 'Buyer', 'Quantity (tons)', 'Price per Ton', 'Total Amount (₹)']
            for col, header in enumerate(headers, 1):
                cell = ws_sales.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            row = 2
            for sale in sales:
                ws_sales.cell(row=row, column=1, value=sale.sale_date.strftime('%Y-%m-%d'))
                ws_sales.cell(row=row, column=2, value=sale.crop_produce.crop.name)
                ws_sales.cell(row=row, column=3, value=sale.buyer)
                ws_sales.cell(row=row, column=4, value=float(sale.quantity))
                ws_sales.cell(row=row, column=5, value=float(sale.price_per_ton))
                ws_sales.cell(row=row, column=6, value=float(sale.total_amount))
                row += 1
        
        # If no sheets were created, create a default one
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet("No Data")
            ws.cell(row=1, column=1, value="No transactions found for the selected period.")
        
        # Auto-adjust column widths for all sheets
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{farm.name}_Finance_Ledger_{start_date}_{end_date}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class ItemCreateView(LoginRequiredMixin, CreateView):
    model = Item
    form_class = ItemForm
    template_name = 'farm/generic_form.html'
    success_url = '/farmlands/'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_type'] = 'Item'
        return context


class CropCreateView(LoginRequiredMixin, CreateView):
    model = Crop
    form_class = CropForm
    template_name = 'farm/generic_form.html'
    success_url = '/farmlands/'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_type'] = 'Crop'
        return context
    
    def dispatch(self, request, *args, **kwargs):
        # Only allow OPERATION_DIRECTOR and ADMIN
        if not request.user.is_superuser:
            if not (hasattr(request.user, 'staff') and 
                   request.user.staff.role.name in ['OPERATION_DIRECTOR', 'ADMIN']):
                from django.core.exceptions import PermissionDenied
                raise PermissionDenied("Only Operation Directors can add crops.")
        return super().dispatch(request, *args, **kwargs)


class CropProduceCreateView(LoginRequiredMixin, FarmContextMixin, CreateView):
    model = CropProduce
    fields = ['crop', 'plot', 'harvest_date', 'quantity', 'storage_location']
    template_name = 'farm/generic_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_type'] = 'Crop Produce'
        return context
    
    def form_valid(self, form):
        form.instance.farm = self.farm
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('farm:farmland_detail', kwargs={'pk': self.farm.pk})
    
    def dispatch(self, request, *args, **kwargs):
        # Only allow FARM_MANAGER and above
        if not request.user.is_superuser:
            if not (hasattr(request.user, 'staff') and 
                   request.user.staff.role.name in ['FARM_MANAGER', 'OPERATION_DIRECTOR', 'ADMIN']):
                from django.core.exceptions import PermissionDenied
                raise PermissionDenied("Only Farm Managers can add crop produce.")
        return super().dispatch(request, *args, **kwargs)